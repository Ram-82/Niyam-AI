"""
Export + CA Workflow Layer — filing-ready output generation.

Converts processed Niyam AI data into clean formats for:
    1. JSON (primary API output, structured for systems)
    2. Excel (CA-friendly, multi-sheet workbook)
    3. CSV (lightweight, tool-import-ready)

Also computes "Ready to File" status and CA handoff readiness.

Pipeline position:
    Invoices + Compliance + ITC → **Export Service** → JSON / Excel / CSV
"""

import csv
import io
import json
import logging
from datetime import date, datetime
from enum import Enum
from typing import List, Dict, Optional, Any

logger = logging.getLogger(__name__)


# ============================================================
# Ready-to-File Logic
# ============================================================

# Blocking severity levels — if ANY flag has these, NOT ready to file
_BLOCKING_SEVERITIES = {"critical", "error"}

# Minimum average confidence for filing readiness
_MIN_FILING_CONFIDENCE = 60


def assess_filing_readiness(
    invoices: List[dict],
    compliance_flags: List[dict],
    itc_results: List[dict],
) -> dict:
    """
    Determine if the dataset is ready for filing.

    Returns:
        {
            "ready_to_file": bool,
            "blocking_issues": [...],
            "warnings": [...],
            "clean_invoice_count": int,
            "total_invoice_count": int,
            "clean_rate": float (0-100),
        }
    """
    blocking = []
    warnings = []

    # ---- Check 1: Critical/error compliance flags ----
    for f in compliance_flags:
        sev = f.get("severity", "info")
        sev_str = sev.value if hasattr(sev, "value") else str(sev).split(".")[-1].lower()

        if sev_str in _BLOCKING_SEVERITIES:
            rule_id = f.get("rule_id", "unknown")
            msg = f.get("message", rule_id)
            action = f.get("action_required", "")
            blocking.append({
                "source": "compliance",
                "rule_id": rule_id,
                "severity": sev_str,
                "message": msg,
                "action_required": action,
            })
        elif sev_str == "warning":
            warnings.append({
                "source": "compliance",
                "rule_id": f.get("rule_id", "unknown"),
                "message": f.get("message", ""),
            })

    # ---- Check 2: ITC issues ----
    for r in itc_results:
        mt = r.get("match_type", "")
        mt_str = mt.value if hasattr(mt, "value") else str(mt)
        sev = r.get("severity", "none")
        sev_str = sev.value if hasattr(sev, "value") else str(sev)

        if mt_str == "duplicate_claim":
            blocking.append({
                "source": "itc",
                "rule_id": "duplicate_claim",
                "severity": "critical",
                "message": f"Duplicate ITC claim: {r.get('invoice_number', 'unknown')}",
                "action_required": r.get("action_required", "Remove duplicate"),
            })
        elif mt_str == "missing_in_2b":
            itc_risk = float(r.get("itc_at_risk") or 0)
            if itc_risk > 0:
                warnings.append({
                    "source": "itc",
                    "rule_id": "missing_in_2b",
                    "message": f"₹{itc_risk:,.0f} ITC at risk — {r.get('invoice_number', '')} not in GSTR-2B",
                })

    # ---- Check 3: Invoice confidence ----
    confidences = []
    needs_review_count = 0
    clean_count = 0

    for inv in invoices:
        conf = float(inv.get("confidence") or inv.get("confidence_score") or 0)
        confidences.append(conf)
        needs_review = inv.get("needs_review", False)

        if needs_review or conf < _MIN_FILING_CONFIDENCE:
            needs_review_count += 1
        else:
            clean_count += 1

    avg_confidence = sum(confidences) / len(confidences) if confidences else 0

    if avg_confidence < _MIN_FILING_CONFIDENCE and confidences:
        blocking.append({
            "source": "quality",
            "rule_id": "low_average_confidence",
            "severity": "error",
            "message": f"Average extraction confidence {avg_confidence:.0f}% is below {_MIN_FILING_CONFIDENCE}% threshold",
            "action_required": "Review flagged invoices before filing",
        })

    if needs_review_count > 0:
        warnings.append({
            "source": "quality",
            "rule_id": "invoices_need_review",
            "message": f"{needs_review_count} of {len(invoices)} invoices need manual review",
        })

    total = len(invoices)
    clean_rate = (clean_count / total * 100) if total > 0 else 100.0

    return {
        "ready_to_file": len(blocking) == 0,
        "blocking_issues": blocking,
        "warnings": warnings,
        "clean_invoice_count": clean_count,
        "total_invoice_count": total,
        "clean_rate": round(clean_rate, 1),
    }


# ============================================================
# Enum-safe serializer
# ============================================================

def _serialize(obj: Any) -> Any:
    """Recursively convert enums and dates to JSON-safe types."""
    if isinstance(obj, Enum):
        return obj.value
    if isinstance(obj, (date, datetime)):
        return obj.isoformat()
    if isinstance(obj, dict):
        return {k: _serialize(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [_serialize(v) for v in obj]
    return obj


# ============================================================
# Export Filters
# ============================================================

class ExportFilter:
    """Controls which records appear in the export."""

    __slots__ = (
        "clean_only",         # exclude needs_review invoices
        "exclude_high_risk",  # exclude critical/error flagged items
        "include_flagged",    # include items with any flag
        "min_confidence",     # minimum confidence threshold
    )

    def __init__(
        self,
        clean_only: bool = False,
        exclude_high_risk: bool = False,
        include_flagged: bool = True,
        min_confidence: int = 0,
    ):
        self.clean_only = clean_only
        self.exclude_high_risk = exclude_high_risk
        self.include_flagged = include_flagged
        self.min_confidence = min_confidence

    def filter_invoices(self, invoices: List[dict]) -> List[dict]:
        """Apply filters to invoice list."""
        result = []
        for inv in invoices:
            conf = float(inv.get("confidence") or inv.get("confidence_score") or 100)
            needs_review = inv.get("needs_review", False)

            if self.clean_only and needs_review:
                continue
            if conf < self.min_confidence:
                continue

            result.append(inv)
        return result

    def filter_itc(self, itc_results: List[dict]) -> List[dict]:
        """Apply filters to ITC results."""
        result = []
        for r in itc_results:
            sev = r.get("severity", "none")
            sev_str = sev.value if hasattr(sev, "value") else str(sev)

            if self.exclude_high_risk and sev_str in ("critical", "high"):
                continue
            if not self.include_flagged and r.get("risk_flag"):
                continue

            result.append(r)
        return result

    def filter_flags(self, flags: List[dict]) -> List[dict]:
        """Apply filters to compliance flags."""
        if not self.include_flagged:
            return []
        result = []
        for f in flags:
            sev = f.get("severity", "info")
            sev_str = sev.value if hasattr(sev, "value") else str(sev).split(".")[-1].lower()
            if self.exclude_high_risk and sev_str in ("critical", "error"):
                continue
            result.append(f)
        return result


# ============================================================
# JSON Export
# ============================================================

def export_json(
    business: dict,
    period: str,
    invoices: List[dict],
    compliance_flags: List[dict],
    itc_results: List[dict],
    itc_financials: Optional[dict],
    filing_readiness: dict,
    export_filter: ExportFilter = None,
) -> dict:
    """
    Generate structured JSON export — primary format.

    This is the canonical representation of all processed data.
    """
    filt = export_filter or ExportFilter()

    filtered_inv = filt.filter_invoices(invoices)
    filtered_itc = filt.filter_itc(itc_results)
    filtered_flags = filt.filter_flags(compliance_flags)

    payload = {
        "export_format": "json",
        "generated_at": datetime.utcnow().isoformat() + "Z",
        "disclaimer": (
            "Niyam AI provides compliance insights for informational purposes only. "
            "Final filing responsibility lies with the user and their Chartered Accountant."
        ),
        "business": _serialize(business),
        "period": period,
        "filing_readiness": _serialize(filing_readiness),
        "invoices": {
            "count": len(filtered_inv),
            "records": _serialize(filtered_inv),
        },
        "itc_summary": _serialize(itc_financials or {}),
        "itc_matching": {
            "count": len(filtered_itc),
            "records": _serialize(filtered_itc),
        },
        "compliance_summary": {
            "total_flags": len(filtered_flags),
            "flags": _serialize(filtered_flags),
        },
    }

    return payload


# ============================================================
# Excel Export
# ============================================================

def export_excel(
    business: dict,
    period: str,
    invoices: List[dict],
    compliance_flags: List[dict],
    itc_results: List[dict],
    itc_financials: Optional[dict],
    filing_readiness: dict,
    export_filter: ExportFilter = None,
) -> bytes:
    """
    Generate multi-sheet Excel workbook — CA-friendly format.

    Sheets:
        1. Summary     — business info, filing readiness, financial totals
        2. Invoices    — all invoice data in tabular form
        3. ITC Matching — match results with severity/action
        4. Compliance  — compliance flags with actions
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    filt = export_filter or ExportFilter()
    filtered_inv = filt.filter_invoices(invoices)
    filtered_itc = filt.filter_itc(itc_results)
    filtered_flags = filt.filter_flags(compliance_flags)

    wb = Workbook()

    # ---- Styles ----
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    title_font = Font(bold=True, size=14)
    good_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    bad_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def _write_header(ws, row, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

    def _enum_val(v):
        return v.value if hasattr(v, "value") else str(v) if v is not None else ""

    # ============================================================
    # Sheet 1: Summary
    # ============================================================
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.column_dimensions["A"].width = 30
    ws_summary.column_dimensions["B"].width = 40

    ws_summary.cell(row=1, column=1, value="Niyam AI — Export Report").font = title_font
    ws_summary.cell(row=2, column=1, value=f"Period: {period}")
    ws_summary.cell(row=3, column=1, value=f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}")

    # Business info
    ws_summary.cell(row=5, column=1, value="Business").font = header_font
    row = 6
    for key in ("legal_name", "trade_name", "gstin", "pan", "business_type"):
        val = business.get(key, "")
        ws_summary.cell(row=row, column=1, value=key.replace("_", " ").title())
        ws_summary.cell(row=row, column=2, value=str(val) if val else "")
        row += 1

    # Filing readiness
    row += 1
    ws_summary.cell(row=row, column=1, value="Filing Readiness").font = header_font
    row += 1
    ready = filing_readiness.get("ready_to_file", False)
    cell = ws_summary.cell(row=row, column=1, value="Ready to File")
    ws_summary.cell(row=row, column=2, value="YES" if ready else "NO")
    ws_summary.cell(row=row, column=2).fill = good_fill if ready else bad_fill
    row += 1
    ws_summary.cell(row=row, column=1, value="Clean Invoices")
    ws_summary.cell(row=row, column=2, value=f"{filing_readiness.get('clean_invoice_count', 0)} / {filing_readiness.get('total_invoice_count', 0)}")
    row += 1
    ws_summary.cell(row=row, column=1, value="Clean Rate")
    ws_summary.cell(row=row, column=2, value=f"{filing_readiness.get('clean_rate', 0):.1f}%")
    row += 1
    ws_summary.cell(row=row, column=1, value="Blocking Issues")
    ws_summary.cell(row=row, column=2, value=str(len(filing_readiness.get("blocking_issues", []))))

    # Blocking issues list
    for issue in filing_readiness.get("blocking_issues", []):
        row += 1
        cell_msg = ws_summary.cell(row=row, column=2, value=issue.get("message", ""))
        cell_msg.fill = bad_fill

    # ITC financials
    row += 2
    ws_summary.cell(row=row, column=1, value="ITC Financial Summary").font = header_font
    row += 1
    fin = itc_financials or {}
    for label, key in [
        ("Total ITC Available", "total_itc_available"),
        ("Total ITC Claimed", "total_itc_claimed"),
        ("Total ITC at Risk", "total_itc_at_risk"),
        ("Recoverable ITC", "recoverable_itc"),
        ("Net ITC Position", "net_itc_position"),
        ("Utilization Rate", "utilization_rate"),
    ]:
        val = fin.get(key, 0)
        ws_summary.cell(row=row, column=1, value=label)
        if key == "utilization_rate":
            ws_summary.cell(row=row, column=2, value=f"{val:.1f}%")
        else:
            ws_summary.cell(row=row, column=2, value=round(float(val), 2))
            ws_summary.cell(row=row, column=2).number_format = '#,##0.00'
        row += 1

    # ============================================================
    # Sheet 2: Invoices
    # ============================================================
    ws_inv = wb.create_sheet("Invoices")

    inv_headers = [
        "Invoice Number", "Invoice Date", "Vendor Name", "Vendor GSTIN",
        "Taxable Value", "CGST", "SGST", "IGST", "Total Amount",
        "Confidence", "Needs Review", "Review Notes",
    ]
    _write_header(ws_inv, 1, inv_headers)

    for i, inv in enumerate(filtered_inv, 2):
        ws_inv.cell(row=i, column=1, value=inv.get("invoice_number") or "")
        ws_inv.cell(row=i, column=2, value=inv.get("invoice_date") or "")
        ws_inv.cell(row=i, column=3, value=inv.get("vendor_name") or "")
        ws_inv.cell(row=i, column=4, value=inv.get("vendor_gstin") or inv.get("gstin") or "")
        ws_inv.cell(row=i, column=5, value=float(inv.get("taxable_value") or inv.get("taxable_amount") or 0)).number_format = '#,##0.00'
        ws_inv.cell(row=i, column=6, value=float(inv.get("cgst") or 0)).number_format = '#,##0.00'
        ws_inv.cell(row=i, column=7, value=float(inv.get("sgst") or 0)).number_format = '#,##0.00'
        ws_inv.cell(row=i, column=8, value=float(inv.get("igst") or 0)).number_format = '#,##0.00'
        ws_inv.cell(row=i, column=9, value=float(inv.get("total_amount") or 0)).number_format = '#,##0.00'
        ws_inv.cell(row=i, column=10, value=float(inv.get("confidence") or inv.get("confidence_score") or 0))
        nr = inv.get("needs_review", False)
        ws_inv.cell(row=i, column=11, value="YES" if nr else "NO")
        if nr:
            ws_inv.cell(row=i, column=11).fill = warn_fill
        rn = inv.get("review_notes") or ""
        if isinstance(rn, list):
            rn = ", ".join(rn)
        ws_inv.cell(row=i, column=12, value=rn)

        for col in range(1, len(inv_headers) + 1):
            ws_inv.cell(row=i, column=col).border = thin_border

    # Auto-width
    for col_idx in range(1, len(inv_headers) + 1):
        ws_inv.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "A"].width = 16

    # ============================================================
    # Sheet 3: ITC Matching
    # ============================================================
    ws_itc = wb.create_sheet("ITC Matching")

    itc_headers = [
        "Invoice Number", "Vendor GSTIN", "Match Type", "Severity",
        "Action Type", "Eligible ITC", "Claimed ITC", "ITC at Risk",
        "Recovery Priority", "Confidence", "Action Required", "Due Date",
    ]
    _write_header(ws_itc, 1, itc_headers)

    severity_fills = {
        "critical": bad_fill,
        "high": bad_fill,
        "error": bad_fill,
        "medium": warn_fill,
        "warning": warn_fill,
        "low": PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
        "none": good_fill,
    }

    for i, r in enumerate(filtered_itc, 2):
        ws_itc.cell(row=i, column=1, value=r.get("invoice_number") or "")
        ws_itc.cell(row=i, column=2, value=r.get("vendor_gstin") or "")
        ws_itc.cell(row=i, column=3, value=_enum_val(r.get("match_type")))
        sev_val = _enum_val(r.get("severity"))
        ws_itc.cell(row=i, column=4, value=sev_val)
        ws_itc.cell(row=i, column=4).fill = severity_fills.get(sev_val, PatternFill())
        ws_itc.cell(row=i, column=5, value=_enum_val(r.get("action_type")))
        ws_itc.cell(row=i, column=6, value=float(r.get("eligible_itc") or 0)).number_format = '#,##0.00'
        ws_itc.cell(row=i, column=7, value=float(r.get("claimed_itc") or 0)).number_format = '#,##0.00'
        ws_itc.cell(row=i, column=8, value=float(r.get("itc_at_risk") or 0)).number_format = '#,##0.00'
        ws_itc.cell(row=i, column=9, value=_enum_val(r.get("recovery_priority")))
        ws_itc.cell(row=i, column=10, value=int(r.get("confidence_score") or 0))
        ws_itc.cell(row=i, column=11, value=r.get("action_required") or "")
        ws_itc.cell(row=i, column=12, value=r.get("due_date") or "")

        for col in range(1, len(itc_headers) + 1):
            ws_itc.cell(row=i, column=col).border = thin_border

    for col_idx in range(1, len(itc_headers) + 1):
        ws_itc.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "A"].width = 18

    # ============================================================
    # Sheet 4: Compliance Flags
    # ============================================================
    ws_flags = wb.create_sheet("Compliance Flags")

    flag_headers = [
        "Rule ID", "Category", "Severity", "Message",
        "Action Required", "Impact (₹)", "Due Date",
    ]
    _write_header(ws_flags, 1, flag_headers)

    for i, f in enumerate(filtered_flags, 2):
        ws_flags.cell(row=i, column=1, value=f.get("rule_id") or "")
        ws_flags.cell(row=i, column=2, value=_enum_val(f.get("category")))
        sev_val = _enum_val(f.get("severity"))
        ws_flags.cell(row=i, column=3, value=sev_val)
        ws_flags.cell(row=i, column=3).fill = severity_fills.get(sev_val, PatternFill())
        ws_flags.cell(row=i, column=4, value=f.get("message") or "")
        ws_flags.cell(row=i, column=5, value=f.get("action_required") or "")
        ws_flags.cell(row=i, column=6, value=float(f.get("impact_amount") or 0)).number_format = '#,##0.00'
        ws_flags.cell(row=i, column=7, value=f.get("due_date") or "")

        for col in range(1, len(flag_headers) + 1):
            ws_flags.cell(row=i, column=col).border = thin_border

    for col_idx in range(1, len(flag_headers) + 1):
        ws_flags.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "A"].width = 22

    # Add disclaimer to Summary sheet footer
    disclaimer_row = ws_summary.max_row + 3
    disc_cell = ws_summary.cell(
        row=disclaimer_row, column=1,
        value=(
            "DISCLAIMER: Niyam AI provides compliance insights for informational purposes only. "
            "Final filing responsibility lies with the user and their Chartered Accountant."
        ),
    )
    from openpyxl.styles import Font as _Font
    disc_cell.font = _Font(italic=True, size=9, color="808080")
    ws_summary.merge_cells(
        start_row=disclaimer_row, start_column=1,
        end_row=disclaimer_row, end_column=4,
    )

    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ============================================================
# CSV Export
# ============================================================

def export_csv(
    invoices: List[dict],
    itc_results: List[dict],
    compliance_flags: List[dict],
    export_filter: ExportFilter = None,
) -> Dict[str, str]:
    """
    Generate CSV strings for each data type.

    Returns dict of {filename: csv_string}:
        invoices.csv, itc_matching.csv, compliance_flags.csv
    """
    filt = export_filter or ExportFilter()
    filtered_inv = filt.filter_invoices(invoices)
    filtered_itc = filt.filter_itc(itc_results)
    filtered_flags = filt.filter_flags(compliance_flags)

    csvs = {}

    # ---- Invoices CSV ----
    inv_fields = [
        "invoice_number", "invoice_date", "vendor_name", "vendor_gstin",
        "taxable_value", "cgst", "sgst", "igst", "total_amount",
        "confidence", "needs_review",
    ]
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=inv_fields, extrasaction="ignore")
    writer.writeheader()
    for inv in filtered_inv:
        row = {}
        for f in inv_fields:
            val = inv.get(f)
            if f == "vendor_gstin":
                val = inv.get("vendor_gstin") or inv.get("gstin") or ""
            if f == "taxable_value":
                val = inv.get("taxable_value") or inv.get("taxable_amount") or 0
            if f == "confidence":
                val = inv.get("confidence") or inv.get("confidence_score") or 0
            if hasattr(val, "value"):
                val = val.value
            row[f] = val
        writer.writerow(row)
    csvs["invoices.csv"] = buf.getvalue()

    # ---- ITC Matching CSV ----
    itc_fields = [
        "invoice_number", "vendor_gstin", "match_type", "severity",
        "action_type", "eligible_itc", "claimed_itc", "itc_at_risk",
        "recovery_priority", "confidence_score", "action_required", "due_date",
    ]
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=itc_fields, extrasaction="ignore")
    writer.writeheader()
    for r in filtered_itc:
        row = {}
        for f in itc_fields:
            val = r.get(f)
            if hasattr(val, "value"):
                val = val.value
            row[f] = val if val is not None else ""
        writer.writerow(row)
    csvs["itc_matching.csv"] = buf.getvalue()

    # ---- Compliance Flags CSV ----
    flag_fields = [
        "rule_id", "category", "severity", "message",
        "action_required", "impact_amount", "due_date",
    ]
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=flag_fields, extrasaction="ignore")
    writer.writeheader()
    for f in filtered_flags:
        row = {}
        for field in flag_fields:
            val = f.get(field)
            if hasattr(val, "value"):
                val = val.value
            row[field] = val if val is not None else ""
        writer.writerow(row)
    csvs["compliance_flags.csv"] = buf.getvalue()

    return csvs


# ============================================================
# Export Service (main orchestrator)
# ============================================================

class ExportService:
    """
    Orchestrates export generation across all formats.

    Usage:
        service = ExportService()
        result = service.export(
            format="json",
            business=business,
            period="Mar 2026",
            invoices=invoices,
            compliance_flags=flags,
            itc_results=itc_results,
            itc_financials=itc_financials,
        )
    """

    def export(
        self,
        format: str,
        business: dict,
        period: str,
        invoices: List[dict],
        compliance_flags: List[dict] = None,
        itc_results: List[dict] = None,
        itc_financials: Optional[dict] = None,
        clean_only: bool = False,
        exclude_high_risk: bool = False,
        include_flagged: bool = True,
        min_confidence: int = 0,
    ) -> dict:
        """
        Generate export in the specified format.

        Args:
            format: "json", "excel", or "csv"
            business: business info dict
            period: filing period string
            invoices: normalized invoice list
            compliance_flags: Rules Engine flags
            itc_results: ITC match results
            itc_financials: ITC financial summary
            clean_only: only include non-flagged invoices
            exclude_high_risk: exclude critical/error items
            include_flagged: include flagged items (default True)
            min_confidence: minimum confidence filter

        Returns:
            {
                "format": str,
                "filing_readiness": dict,
                "data": dict|bytes|dict[str,str] depending on format
            }
        """
        flags = compliance_flags or []
        itc = itc_results or []

        export_filter = ExportFilter(
            clean_only=clean_only,
            exclude_high_risk=exclude_high_risk,
            include_flagged=include_flagged,
            min_confidence=min_confidence,
        )

        # Always compute filing readiness (unfiltered)
        readiness = assess_filing_readiness(invoices, flags, itc)

        if format == "json":
            data = export_json(
                business=business, period=period, invoices=invoices,
                compliance_flags=flags, itc_results=itc,
                itc_financials=itc_financials, filing_readiness=readiness,
                export_filter=export_filter,
            )
            return {"format": "json", "filing_readiness": readiness, "data": data}

        elif format == "excel":
            data = export_excel(
                business=business, period=period, invoices=invoices,
                compliance_flags=flags, itc_results=itc,
                itc_financials=itc_financials, filing_readiness=readiness,
                export_filter=export_filter,
            )
            return {"format": "excel", "filing_readiness": readiness, "data": data}

        elif format == "csv":
            data = export_csv(
                invoices=invoices, itc_results=itc,
                compliance_flags=flags, export_filter=export_filter,
            )
            return {"format": "csv", "filing_readiness": readiness, "data": data}

        else:
            raise ValueError(f"Unsupported format: {format}. Use json, excel, or csv.")
